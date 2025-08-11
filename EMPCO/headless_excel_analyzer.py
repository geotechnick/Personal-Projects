#!/usr/bin/env python3
"""
Headless Excel Analyzer
Alternative implementation using openpyxl for completely headless Excel processing
without requiring Excel installation or visible interface.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from pathlib import Path
from typing import Dict, Any
import logging


class HeadlessExcelAnalyzer:
    """
    Headless Excel analyzer using openpyxl - no Excel installation required
    
    Note: This approach can read/write Excel files but cannot execute complex formulas.
    For formula execution, xlwings with hidden Excel is still needed.
    """
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.input_sheet = None
        self.calc_sheet = None
        
    def load_workbook(self):
        """Load Excel workbook using openpyxl"""
        try:
            self.workbook = load_workbook(self.excel_path, data_only=False)  # Include formulas
            self.input_sheet = self.workbook['Input&Summary']
            self.calc_sheet = self.workbook['Calcs']
            return True
        except Exception as e:
            logging.error(f"Failed to load workbook: {e}")
            return False
    
    def read_current_values(self) -> Dict[str, Any]:
        """Read current values from the spreadsheet"""
        if not self.workbook:
            self.load_workbook()
            
        try:
            # Read key input values
            inputs = {
                'pipe_od': self.input_sheet['C3'].value,
                'pipe_wt': self.input_sheet['C4'].value, 
                'pipe_smys': self.input_sheet['C5'].value,
                'pipe_doc': self.input_sheet['C6'].value,
                'pipe_length': self.input_sheet['C7'].value,
                'internal_pressure': self.input_sheet['C9'].value,
                'friction_angle': self.input_sheet['F3'].value,
                'cohesion': self.input_sheet['F4'].value,
                'unit_weight': self.input_sheet['F5'].value,
                'pgd_direction': self.input_sheet['F6'].value
            }
            
            # Read current calculated outputs
            outputs = {
                'longitudinal_force': self.input_sheet['C13'].value,
                'axial_stress': self.input_sheet['C14'].value,
                'remaining_allowable_stress': self.input_sheet['C15'].value,
                'allowable_length': self.input_sheet['C16'].value,
                'exceeds_allowable': self.input_sheet['C17'].value
            }
            
            return {'inputs': inputs, 'outputs': outputs}
            
        except Exception as e:
            logging.error(f"Failed to read values: {e}")
            return {}
    
    def create_lookup_table(self) -> pd.DataFrame:
        """
        Create a lookup table from existing Excel calculations
        This approach pre-calculates results for common parameter ranges
        """
        if not self.workbook:
            self.load_workbook()
            
        # This would involve:
        # 1. Reading the Excel formulas
        # 2. Creating parameter ranges
        # 3. Building lookup table for interpolation
        
        # Placeholder implementation
        return pd.DataFrame()
    
    def interpolate_results(self, pipe_od: float, pipe_wt: float, 
                          friction_angle: float, cohesion: float) -> Dict[str, float]:
        """
        Interpolate results from lookup table instead of running Excel formulas
        """
        # This would implement interpolation logic
        # For now, return placeholder values
        return {
            'longitudinal_force': 1500.0,
            'axial_stress': 300.0,
            'remaining_allowable_stress': 12000.0,
            'allowable_length': 200.0,
            'exceeds_allowable': False
        }


class HybridExcelAnalyzer:
    """
    Hybrid approach: Use xlwings in background mode for formula execution
    with minimal interface visibility
    """
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.app = None
        self.wb = None
        
    def __enter__(self):
        """Context manager entry - start hidden Excel"""
        try:
            import xlwings as xw
            
            # Start Excel in completely hidden mode
            self.app = xw.App(visible=False, add_book=False)
            self.app.display_alerts = False
            self.app.screen_updating = False
            
            # Open workbook
            self.wb = self.app.books.open(str(self.excel_path))
            
            return self
            
        except Exception as e:
            logging.error(f"Failed to start hidden Excel: {e}")
            raise
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - clean up Excel"""
        try:
            if self.wb:
                self.wb.close()
            if self.app:
                self.app.quit()
        except:
            pass
    
    def calculate_soil_springs(self, pipe_config: Dict, soil_config: Dict) -> Dict[str, float]:
        """Calculate soil springs using hidden Excel"""
        try:
            input_sheet = self.wb.sheets['Input&Summary']
            
            # Update inputs
            input_sheet.range('C3').value = pipe_config.get('pipe_od', 16)
            input_sheet.range('C4').value = pipe_config.get('pipe_wt', 0.375)
            input_sheet.range('C6').value = pipe_config.get('pipe_doc', 10)
            input_sheet.range('C7').value = pipe_config.get('pipe_length', 10)
            input_sheet.range('C9').value = pipe_config.get('internal_pressure', 1440)
            
            input_sheet.range('F3').value = soil_config.get('friction_angle', 30)
            input_sheet.range('F4').value = soil_config.get('cohesion', 100)
            input_sheet.range('F5').value = soil_config.get('unit_weight', 125)
            input_sheet.range('F6').value = pipe_config.get('pgd_direction', 'Parallel')
            
            # Force calculation
            self.app.calculate()
            
            # Read results
            results = {
                'longitudinal_force': input_sheet.range('C13').value or 0,
                'axial_stress': input_sheet.range('C14').value or 0,
                'remaining_allowable_stress': input_sheet.range('C15').value or 0,
                'allowable_length': input_sheet.range('C16').value or 0,
                'exceeds_allowable': input_sheet.range('C17').value == "Exceeds"
            }
            
            return results
            
        except Exception as e:
            logging.error(f"Excel calculation failed: {e}")
            return {}


def demonstrate_headless_excel():
    """Demonstrate headless Excel processing"""
    
    print("=== Headless Excel Analysis Demo ===")
    
    excel_path = "Soil Springs_2024.xlsx"
    
    # Method 1: Completely headless (read-only)
    print("\n1. Headless read-only analysis:")
    headless = HeadlessExcelAnalyzer(excel_path)
    if headless.load_workbook():
        current_values = headless.read_current_values()
        print(f"Current pipe OD: {current_values.get('inputs', {}).get('pipe_od', 'N/A')}")
        print(f"Current longitudinal force: {current_values.get('outputs', {}).get('longitudinal_force', 'N/A')}")
    
    # Method 2: Hidden Excel with calculations
    print("\n2. Hidden Excel with calculations:")
    try:
        with HybridExcelAnalyzer(excel_path) as analyzer:
            pipe_config = {
                'pipe_od': 20,
                'pipe_wt': 0.5,
                'pipe_doc': 8,
                'pipe_length': 15,
                'internal_pressure': 1200,
                'pgd_direction': 'Parallel'
            }
            
            soil_config = {
                'friction_angle': 32,
                'cohesion': 150,
                'unit_weight': 120
            }
            
            results = analyzer.calculate_soil_springs(pipe_config, soil_config)
            print(f"Calculated longitudinal force: {results.get('longitudinal_force', 'Failed')}")
            print(f"Calculated axial stress: {results.get('axial_stress', 'Failed')}")
            
    except ImportError:
        print("xlwings not available - cannot run hidden Excel calculations")
    except Exception as e:
        print(f"Hidden Excel analysis failed: {e}")


if __name__ == "__main__":
    demonstrate_headless_excel()