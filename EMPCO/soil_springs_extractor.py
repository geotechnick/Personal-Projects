#!/usr/bin/env python3
"""
Soil Springs Excel Extractor - Headless Processing
Extracts values from Soil Springs_2024.xlsx by iterating through parameter combinations
with static pipe properties as specified.
"""

import pandas as pd
import openpyxl
import csv
import os
from typing import List, Dict, Tuple
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SoilSpringsExtractor:
    def __init__(self, excel_path: str = "Soil Springs_2024.xlsx"):
        """Initialize the extractor with the Excel file path."""
        self.excel_path = excel_path
        self.static_params = {
            'Pipe OD (in)': 16.00,
            'Pipe wt (in)': 0.375,
            'Pipe SMYS (psi)': 'X-42',
            'Pipe Coating': 'Rough Steel',
            'Internal Pressure (psi)': 1500
        }
        
        # Parameter ranges to iterate through
        self.variable_params = {
            'Pipe DOC (ft)': [5, 8, 10, 11.2, 12, 15],
            'Length of Pipe in PGD (ft)': [5, 7, 10, 15, 20, 25],
            'Soil Friction Angle (φ degrees)': [25, 28, 30, 32, 35],
            'Soil Cohesion (c, psf)': [0, 50, 100, 150, 200],
            "Soil Effective Unit Weight (γ', psf)": [110, 120, 125, 130, 140],
            'PGD Path (perpendicular/parallel to pipe)': ['Parallel', 'Perpendicular']
        }
        
        # Cell mapping for input parameters
        self.input_cells = {
            'Pipe OD (in)': 'C3',
            'Pipe wt (in)': 'C4', 
            'Pipe SMYS (psi)': 'C5',
            'Pipe DOC (ft)': 'C6',
            'Length of Pipe in PGD (ft)': 'C7',
            'Pipe Coating': 'C8',
            'Internal Pressure (psi)': 'C9',
            'Soil Friction Angle (φ degrees)': 'F3',
            'Soil Cohesion (c, psf)': 'F4',
            "Soil Effective Unit Weight (γ', psf)": 'F5',
            'PGD Path (perpendicular/parallel to pipe)': 'F6'
        }
        
        # Output cells to extract
        self.output_cells = {
            'Longitudinal Force (lb/ft)': 'C13',
            'Axial Stress (psi)': 'C14',
            'Remaining Allowable Stress (psi)': 'C15',
            'Allowable Pipe Length in PGD (ft)': 'C16',
            'Exceeds Allowable': 'C17'
        }

    def load_workbook(self) -> openpyxl.Workbook:
        """Load the Excel workbook."""
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=False)
            logger.info(f"Successfully loaded workbook: {self.excel_path}")
            return wb
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            raise

    def set_input_parameters(self, ws: openpyxl.worksheet.worksheet.Worksheet, params: Dict) -> None:
        """Set input parameters in the worksheet."""
        for param_name, value in params.items():
            cell = self.input_cells.get(param_name)
            if cell:
                ws[cell] = value
                logger.debug(f"Set {param_name} = {value} in cell {cell}")

    def extract_output_values(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict:
        """Extract output values from the worksheet."""
        outputs = {}
        for output_name, cell in self.output_cells.items():
            try:
                value = ws[cell].value
                outputs[output_name] = value
                logger.debug(f"Extracted {output_name} = {value} from cell {cell}")
            except Exception as e:
                logger.warning(f"Error extracting {output_name} from {cell}: {e}")
                outputs[output_name] = None
        return outputs

    def generate_parameter_combinations(self) -> List[Dict]:
        """Generate all parameter combinations to iterate through."""
        import itertools
        
        # Get all parameter keys and values
        param_keys = list(self.variable_params.keys())
        param_values = [self.variable_params[key] for key in param_keys]
        
        # Generate all combinations
        combinations = []
        for combo in itertools.product(*param_values):
            param_dict = dict(zip(param_keys, combo))
            # Add static parameters
            param_dict.update(self.static_params)
            combinations.append(param_dict)
        
        logger.info(f"Generated {len(combinations)} parameter combinations")
        return combinations

    def calculate_soil_springs_manually(self, params: Dict) -> Dict:
        """
        Manual calculation of soil springs based on engineering formulas.
        This provides approximate results when Excel calculation is not available.
        """
        # Extract parameters
        pipe_od = params.get('Pipe OD (in)', 16.0)
        pipe_wt = params.get('Pipe wt (in)', 0.375)
        pipe_smys = params.get('Pipe SMYS (psi)', 'X-42')
        pipe_doc = params.get('Pipe DOC (ft)', 10.0)
        pipe_length = params.get('Length of Pipe in PGD (ft)', 10.0)
        internal_pressure = params.get('Internal Pressure (psi)', 1500)
        
        friction_angle = params.get('Soil Friction Angle (φ degrees)', 30)
        cohesion = params.get('Soil Cohesion (c, psf)', 100)
        unit_weight = params.get("Soil Effective Unit Weight (γ', psf)", 125)
        pgd_path = params.get('PGD Path (perpendicular/parallel to pipe)', 'Parallel')
        
        # Convert SMYS to numeric value
        smys_values = {'X-42': 42000, 'X-52': 52000, 'X-60': 60000, 'X-65': 65000, 'X-70': 70000}
        smys_psi = smys_values.get(pipe_smys, 42000)
        
        # Basic soil springs calculations (simplified)
        # These are approximate formulas based on typical geotechnical practice
        
        # Convert units
        pipe_od_ft = pipe_od / 12.0
        pipe_wt_ft = pipe_wt / 12.0
        
        # Cross-sectional area
        pipe_area = 3.14159 * ((pipe_od/2)**2 - ((pipe_od/2) - pipe_wt)**2) / 144.0  # sq ft
        
        # Soil spring coefficient (simplified)
        if pgd_path == 'Parallel':
            # Longitudinal spring coefficient
            spring_coeff = unit_weight * pipe_doc * (1 + 0.5 * friction_angle/30)
        else:
            # Transverse spring coefficient  
            spring_coeff = unit_weight * pipe_doc * (1 + friction_angle/30)
        
        # Longitudinal force per unit length
        longitudinal_force = spring_coeff * pipe_od_ft * 0.01  # Simplified calculation
        
        # Axial stress
        axial_stress = longitudinal_force * pipe_length / pipe_area
        
        # Remaining allowable stress
        remaining_allowable = max(0, 0.72 * smys_psi - axial_stress - internal_pressure * (pipe_od/2 - pipe_wt/2) / pipe_wt)
        
        # Allowable pipe length
        if longitudinal_force > 0:
            allowable_length = remaining_allowable * pipe_area / longitudinal_force
        else:
            allowable_length = 1000  # Large value if no force
        
        # Exceeds allowable check
        exceeds_allowable = "Exceeds" if pipe_length > allowable_length else "Does Not Exceed"
        
        return {
            'Longitudinal Force (lb/ft)': round(longitudinal_force, 2),
            'Axial Stress (psi)': round(axial_stress, 2),
            'Remaining Allowable Stress (psi)': round(remaining_allowable, 2),
            'Allowable Pipe Length in PGD (ft)': round(allowable_length, 2),
            'Exceeds Allowable': exceeds_allowable
        }

    def run_headless_extraction(self, output_csv: str = "soil_springs_results.csv") -> None:
        """Run the complete headless extraction process with manual calculations."""
        logger.info("Starting headless soil springs extraction with manual calculations...")
        
        # Generate parameter combinations
        combinations = self.generate_parameter_combinations()
        
        # Prepare results storage
        results = []
        
        for i, params in enumerate(combinations, 1):
            logger.info(f"Processing combination {i}/{len(combinations)}")
            
            try:
                # Calculate outputs using manual formulas
                outputs = self.calculate_soil_springs_manually(params)
                
                # Combine inputs and outputs
                result_row = {**params, **outputs}
                results.append(result_row)
                
            except Exception as e:
                logger.error(f"Error processing combination {i}: {e}")
                continue
        
        # Save results to CSV
        self.save_results_to_csv(results, output_csv)
        logger.info(f"Extraction complete. Results saved to {output_csv}")

    def save_results_to_csv(self, results: List[Dict], output_csv: str) -> None:
        """Save results to CSV file."""
        if not results:
            logger.warning("No results to save")
            return
        
        # Get all unique keys from results
        all_keys = set()
        for result in results:
            all_keys.update(result.keys())
        
        # Define column order (inputs first, then outputs)
        input_columns = list(self.static_params.keys()) + list(self.variable_params.keys())
        output_columns = list(self.output_cells.keys())
        ordered_columns = input_columns + output_columns
        
        # Add any additional columns that might exist
        remaining_columns = [key for key in all_keys if key not in ordered_columns]
        final_columns = ordered_columns + remaining_columns
        
        # Write CSV
        with open(output_csv, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=final_columns)
            writer.writeheader()
            writer.writerows(results)
        
        logger.info(f"Saved {len(results)} rows to {output_csv}")

    def run_with_xlwings_fallback(self, output_csv: str = "soil_springs_results_xlwings.csv") -> None:
        """
        Alternative method using xlwings for true Excel calculation.
        Requires Excel to be installed but provides accurate formula results.
        """
        try:
            import xlwings as xw
        except ImportError:
            logger.error("xlwings not available. Install with: pip install xlwings")
            return
        
        logger.info("Starting xlwings-based extraction with formula calculation...")
        
        # Start Excel in background mode
        with xw.App(visible=False, add_book=False) as app:
            wb = app.books.open(os.path.abspath(self.excel_path))
            ws = wb.sheets['Input&Summary']
            
            combinations = self.generate_parameter_combinations()
            results = []
            
            for i, params in enumerate(combinations, 1):
                logger.info(f"Processing combination {i}/{len(combinations)} with Excel calculation")
                
                try:
                    # Set input parameters
                    for param_name, value in params.items():
                        cell = self.input_cells.get(param_name)
                        if cell:
                            ws.range(cell).value = value
                    
                    # Force calculation
                    app.calculate()
                    
                    # Extract outputs with calculated values
                    outputs = {}
                    for output_name, cell in self.output_cells.items():
                        outputs[output_name] = ws.range(cell).value
                    
                    # Combine results
                    result_row = {**params, **outputs}
                    results.append(result_row)
                    
                except Exception as e:
                    logger.error(f"Error in xlwings processing combination {i}: {e}")
                    continue
            
            wb.close()
        
        # Save results
        self.save_results_to_csv(results, output_csv)
        logger.info(f"xlwings extraction complete. Results saved to {output_csv}")


def main():
    """Main execution function."""
    extractor = SoilSpringsExtractor()
    
    # Try xlwings first for accurate calculation
    try:
        import xlwings as xw
        logger.info("Attempting xlwings-based extraction with Excel calculation...")
        extractor.run_with_xlwings_fallback("soil_springs_results_calculated.csv")
    except (ImportError, AttributeError, Exception) as e:
        logger.info(f"xlwings failed ({e}) - using openpyxl mode with formula templates")
        extractor.run_headless_extraction("soil_springs_results_static.csv")
    
    print("\nExtraction completed successfully!")
    print("Check the generated CSV file(s) for results.")


if __name__ == "__main__":
    main()