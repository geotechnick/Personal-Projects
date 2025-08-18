#!/usr/bin/env python3
"""
Static Values Iterator - Parameter Combination Generator
Extracts pipe and soil assumptions from Static Values.xlsx and generates
all possible parameter combinations for each soil layer in separate CSV files.
"""

import openpyxl
import pandas as pd
import itertools
import csv
import os
from typing import List, Dict, Tuple, Any
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class StaticValuesIterator:
    def __init__(self, excel_path: str = "Static Values.xlsx"):
        """Initialize the iterator with the Excel file path."""
        self.excel_path = excel_path
        self.pipe_assumptions = {}
        self.soil_layers = []
        
    def load_pipe_assumptions(self) -> Dict[str, Dict[str, Any]]:
        """Load pipe assumptions from the Excel file."""
        logger.info("Loading pipe assumptions...")
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb['Pipe Assumptions']
        
        pipe_assumptions = {}
        
        # Extract pipe parameters with min/max values
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            param_name = ws.cell(row=row, column=1).value  # Column A
            min_value = ws.cell(row=row, column=2).value   # Column B
            max_value = ws.cell(row=row, column=3).value   # Column C
            
            if param_name and param_name.strip() and min_value is not None:
                param_name = param_name.strip().rstrip(':')
                
                # Handle different data types
                if isinstance(min_value, (int, float)) and isinstance(max_value, (int, float)):
                    pipe_assumptions[param_name] = {
                        'min': min_value,
                        'max': max_value,
                        'type': 'numeric'
                    }
                else:
                    # For non-numeric values, treat as categorical
                    pipe_assumptions[param_name] = {
                        'min': str(min_value),
                        'max': str(max_value),
                        'type': 'categorical'
                    }
                
                logger.debug(f"Loaded {param_name}: min={min_value}, max={max_value}")
        
        wb.close()
        self.pipe_assumptions = pipe_assumptions
        logger.info(f"Loaded {len(pipe_assumptions)} pipe parameters")
        return pipe_assumptions
    
    def load_soil_assumptions(self) -> List[Dict[str, Any]]:
        """Load soil layer assumptions from the Excel file."""
        logger.info("Loading soil assumptions...")
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb['Soil Assumptions']
        
        soil_layers = []
        
        # Extract soil layer data
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            soil_name = ws.cell(row=row, column=1).value      # Column A
            soil_type = ws.cell(row=row, column=2).value      # Column B
            unit_weight = ws.cell(row=row, column=3).value    # Column C
            cohesion = ws.cell(row=row, column=4).value       # Column D
            friction_angle = ws.cell(row=row, column=5).value # Column E
            
            if soil_name and soil_name.strip():
                soil_layer = {
                    'name': soil_name.strip(),
                    'type': soil_type.strip() if soil_type else '',
                    'unit_weight': float(unit_weight) if unit_weight else 120.0,
                    'cohesion': float(cohesion) if cohesion else 0.0,
                    'friction_angle': float(friction_angle) if friction_angle else 30.0
                }
                soil_layers.append(soil_layer)
                logger.debug(f"Loaded soil layer: {soil_layer}")
        
        wb.close()
        self.soil_layers = soil_layers
        logger.info(f"Loaded {len(soil_layers)} soil layers")
        return soil_layers
    
    def generate_pipe_parameter_ranges(self) -> Dict[str, List[Any]]:
        """Generate parameter ranges for pipe assumptions."""
        logger.info("Generating pipe parameter ranges...")
        
        parameter_ranges = {}
        
        for param_name, param_data in self.pipe_assumptions.items():
            if param_data['type'] == 'numeric':
                min_val = param_data['min']
                max_val = param_data['max']
                
                # Special handling for different parameters
                if 'DOC' in param_name or 'Length' in param_name:
                    # Generate 1-foot increments for DOC and Length
                    values = list(range(int(min_val), int(max_val) + 1, 1))
                    parameter_ranges[param_name] = values
                    logger.debug(f"{param_name}: {len(values)} values from {min_val} to {max_val}")
                else:
                    # For other numeric parameters, use min and max only
                    if min_val == max_val:
                        parameter_ranges[param_name] = [min_val]
                    else:
                        parameter_ranges[param_name] = [min_val, max_val]
                    logger.debug(f"{param_name}: {parameter_ranges[param_name]}")
            
            else:  # categorical
                # For categorical parameters, collect unique values
                if param_data['min'] == param_data['max']:
                    parameter_ranges[param_name] = [param_data['min']]
                else:
                    parameter_ranges[param_name] = [param_data['min'], param_data['max']]
                logger.debug(f"{param_name}: {parameter_ranges[param_name]}")
        
        return parameter_ranges
    
    def generate_all_combinations(self, pipe_ranges: Dict[str, List[Any]], soil_layer: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate all possible parameter combinations for a given soil layer."""
        logger.info(f"Generating combinations for soil layer: {soil_layer['name']}")
        
        # Get parameter names and their value ranges
        param_names = list(pipe_ranges.keys())
        param_values = [pipe_ranges[name] for name in param_names]
        
        # Generate all combinations
        combinations = []
        for combo in itertools.product(*param_values):
            combination_dict = dict(zip(param_names, combo))
            
            # Add soil layer properties
            combination_dict.update({
                'Soil Name': soil_layer['name'],
                'Soil Type': soil_layer['type'],
                "Soil Effective Unit Weight (γ', psf)": soil_layer['unit_weight'],
                'Soil Cohesion (c, psf)': soil_layer['cohesion'],
                'Soil Friction Angle (φ degrees)': soil_layer['friction_angle']
            })
            
            combinations.append(combination_dict)
        
        logger.info(f"Generated {len(combinations)} combinations for {soil_layer['name']}")
        return combinations
    
    def save_combinations_to_csv(self, combinations: List[Dict[str, Any]], soil_layer_name: str, output_dir: str = "static_values_output") -> str:
        """Save parameter combinations to CSV file."""
        if not combinations:
            logger.warning(f"No combinations to save for {soil_layer_name}")
            return ""
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Create filename
        safe_name = soil_layer_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        csv_filename = os.path.join(output_dir, f"{safe_name}_combinations.csv")
        
        # Define column order
        pipe_columns = list(self.pipe_assumptions.keys())
        soil_columns = ['Soil Name', 'Soil Type', "Soil Effective Unit Weight (γ', psf)", 
                       'Soil Cohesion (c, psf)', 'Soil Friction Angle (φ degrees)']
        all_columns = pipe_columns + soil_columns
        
        # Write CSV
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=all_columns)
            writer.writeheader()
            writer.writerows(combinations)
        
        logger.info(f"Saved {len(combinations)} combinations to {csv_filename}")
        return csv_filename
    
    def run_complete_analysis(self, output_dir: str = "static_values_output") -> List[str]:
        """Run the complete analysis and generate CSV files for all soil layers."""
        logger.info("Starting complete static values analysis...")
        
        # Load data
        self.load_pipe_assumptions()
        self.load_soil_assumptions()
        
        # Generate pipe parameter ranges
        pipe_ranges = self.generate_pipe_parameter_ranges()
        
        # Calculate total combinations for reporting
        total_pipe_combinations = 1
        for param_name, values in pipe_ranges.items():
            total_pipe_combinations *= len(values)
        
        logger.info(f"Total pipe parameter combinations: {total_pipe_combinations}")
        logger.info(f"Number of soil layers: {len(self.soil_layers)}")
        logger.info(f"Total files to generate: {len(self.soil_layers)}")
        
        # Generate CSV files for each soil layer
        generated_files = []
        
        for i, soil_layer in enumerate(self.soil_layers, 1):
            logger.info(f"Processing soil layer {i}/{len(self.soil_layers)}: {soil_layer['name']}")
            
            # Generate all combinations for this soil layer
            combinations = self.generate_all_combinations(pipe_ranges, soil_layer)
            
            # Save to CSV
            csv_file = self.save_combinations_to_csv(combinations, soil_layer['name'], output_dir)
            if csv_file:
                generated_files.append(csv_file)
        
        logger.info(f"Analysis complete! Generated {len(generated_files)} CSV files")
        return generated_files
    
    def print_summary(self) -> None:
        """Print a summary of the loaded data."""
        print("\n" + "="*60)
        print("STATIC VALUES ANALYSIS SUMMARY")
        print("="*60)
        
        print(f"\n📁 Source File: {self.excel_path}")
        
        print(f"\n🔧 PIPE ASSUMPTIONS ({len(self.pipe_assumptions)} parameters):")
        for param_name, param_data in self.pipe_assumptions.items():
            if param_data['type'] == 'numeric':
                print(f"  • {param_name}: {param_data['min']} to {param_data['max']}")
            else:
                print(f"  • {param_name}: {param_data['min']} / {param_data['max']}")
        
        print(f"\n🏔️ SOIL LAYERS ({len(self.soil_layers)} layers):")
        for soil in self.soil_layers:
            print(f"  • {soil['name']} ({soil['type']}): γ'={soil['unit_weight']}, c={soil['cohesion']}, φ={soil['friction_angle']}°")
        
        # Calculate total combinations
        pipe_ranges = self.generate_pipe_parameter_ranges()
        total_combinations = 1
        for param_name, values in pipe_ranges.items():
            total_combinations *= len(values)
        
        print(f"\n📊 ANALYSIS SCOPE:")
        print(f"  • Total pipe parameter combinations: {total_combinations:,}")
        print(f"  • Number of soil layers: {len(self.soil_layers)}")
        print(f"  • Total CSV files to generate: {len(self.soil_layers)}")
        print(f"  • Total data rows across all files: {total_combinations * len(self.soil_layers):,}")


def main():
    """Main execution function."""
    iterator = StaticValuesIterator("Static Values.xlsx")
    
    # Load and display summary
    iterator.load_pipe_assumptions()
    iterator.load_soil_assumptions()
    iterator.print_summary()
    
    # Run complete analysis automatically
    print(f"\n🚀 Starting CSV generation...")
    generated_files = iterator.run_complete_analysis()
    
    print(f"\n✅ ANALYSIS COMPLETE!")
    print(f"Generated {len(generated_files)} CSV files:")
    for file_path in generated_files:
        print(f"  📄 {file_path}")
    
    print(f"\nFiles saved in: static_values_output/")


if __name__ == "__main__":
    main()