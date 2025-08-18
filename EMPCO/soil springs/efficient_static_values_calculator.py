#!/usr/bin/env python3
"""
Efficient Static Values Calculator with Manual Soil Springs Calculations
Uses manual calculation formulas for soil springs analysis to generate results efficiently.
"""

import openpyxl
import pandas as pd
import itertools
import csv
import os
from typing import List, Dict, Tuple, Any
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class EfficientStaticValuesCalculator:
    def __init__(self, static_values_path: str = "Static Values.xlsx"):
        """Initialize the calculator with Excel file path."""
        self.static_values_path = static_values_path
        self.pipe_assumptions = {}
        self.soil_layers = []
    
    def load_pipe_assumptions(self) -> Dict[str, Dict[str, Any]]:
        """Load pipe assumptions from the Static Values Excel file."""
        logger.info("Loading pipe assumptions...")
        
        wb = openpyxl.load_workbook(self.static_values_path, data_only=True)
        ws = wb['Pipe Assumptions']
        
        pipe_assumptions = {}
        
        for row in range(2, ws.max_row + 1):
            param_name = ws.cell(row=row, column=1).value
            min_value = ws.cell(row=row, column=2).value
            max_value = ws.cell(row=row, column=3).value
            
            if param_name and param_name.strip() and min_value is not None:
                param_name = param_name.strip().rstrip(':')
                
                if isinstance(min_value, (int, float)) and isinstance(max_value, (int, float)):
                    pipe_assumptions[param_name] = {
                        'min': min_value,
                        'max': max_value,
                        'type': 'numeric'
                    }
                else:
                    pipe_assumptions[param_name] = {
                        'min': str(min_value),
                        'max': str(max_value),
                        'type': 'categorical'
                    }
        
        wb.close()
        self.pipe_assumptions = pipe_assumptions
        logger.info(f"Loaded {len(pipe_assumptions)} pipe parameters")
        return pipe_assumptions
    
    def load_soil_assumptions(self) -> List[Dict[str, Any]]:
        """Load soil layer assumptions from the Static Values Excel file."""
        logger.info("Loading soil assumptions...")
        
        wb = openpyxl.load_workbook(self.static_values_path, data_only=True)
        ws = wb['Soil Assumptions']
        
        soil_layers = []
        
        for row in range(2, ws.max_row + 1):
            soil_name = ws.cell(row=row, column=1).value
            soil_type = ws.cell(row=row, column=2).value
            unit_weight = ws.cell(row=row, column=3).value
            cohesion = ws.cell(row=row, column=4).value
            friction_angle = ws.cell(row=row, column=5).value
            
            if soil_name and soil_name.strip():
                soil_layer = {
                    'name': soil_name.strip(),
                    'type': soil_type.strip() if soil_type else '',
                    'unit_weight': float(unit_weight) if unit_weight else 120.0,
                    'cohesion': float(cohesion) if cohesion else 0.0,
                    'friction_angle': float(friction_angle) if friction_angle else 30.0
                }
                soil_layers.append(soil_layer)
        
        wb.close()
        self.soil_layers = soil_layers
        logger.info(f"Loaded {len(soil_layers)} soil layers")
        return soil_layers
    
    def generate_pipe_parameter_ranges(self) -> Dict[str, List[Any]]:
        """Generate parameter ranges for pipe assumptions."""
        logger.info("Generating pipe parameter ranges...")
        
        parameter_ranges = {}
        
        for param_name, param_data in self.pipe_assumptions.items():
            if param_data['type'] == 'numeric':
                min_val = param_data['min']
                max_val = param_data['max']
                
                if 'DOC' in param_name or 'Length' in param_name:
                    # Generate 1-foot increments
                    values = list(range(int(min_val), int(max_val) + 1, 1))
                    parameter_ranges[param_name] = values
                else:
                    if min_val == max_val:
                        parameter_ranges[param_name] = [min_val]
                    else:
                        parameter_ranges[param_name] = [min_val, max_val]
            else:
                if param_data['min'] == param_data['max']:
                    parameter_ranges[param_name] = [param_data['min']]
                else:
                    parameter_ranges[param_name] = [param_data['min'], param_data['max']]
        
        return parameter_ranges
    
    def calculate_soil_springs_results(self, pipe_params: Dict, soil_layer: Dict) -> Dict[str, Any]:
        """
        Calculate soil springs results using manual engineering formulas.
        This replicates the Excel calculations for soil springs analysis.
        """
        # Extract parameters
        pipe_od = pipe_params.get('Pipe OD (in)', 16.0)
        pipe_wt = pipe_params.get('Pipe wt (in)', 0.375)
        pipe_smys = pipe_params.get('Pipe SMYS (psi)', 'X-42')
        pipe_doc = pipe_params.get('Pipe DOC (ft)', 10.0)
        pipe_length = pipe_params.get('Length of Pipe in PGD (ft)', 10.0)
        pipe_coating = pipe_params.get('Pipe Coating', 'Rough Steel')
        internal_pressure = pipe_params.get('Internal Pressure (psi)', 1500)
        pgd_path = pipe_params.get('PGD Path (perpendicular/parallel to pipe)', 'Parallel')
        
        friction_angle = soil_layer['friction_angle']
        cohesion = soil_layer['cohesion']
        unit_weight = soil_layer['unit_weight']
        
        # Convert SMYS to numeric value
        smys_values = {'X-42': 42000, 'X-52': 52000, 'X-60': 60000, 'X-65': 65000, 'X-70': 70000}
        smys_psi = smys_values.get(pipe_smys, 42000)
        
        # Pipe geometry calculations
        pipe_od_ft = pipe_od / 12.0
        pipe_wt_ft = pipe_wt / 12.0
        
        # Cross-sectional area of pipe wall (sq inches)
        outer_radius = pipe_od / 2.0
        inner_radius = outer_radius - pipe_wt
        pipe_area_sqin = 3.14159 * (outer_radius**2 - inner_radius**2)
        
        # Advanced soil springs calculations based on engineering practice
        import math
        
        # Passive earth pressure coefficient
        ka = math.tan(math.radians(45 - friction_angle/2))**2  # Active
        kp = math.tan(math.radians(45 + friction_angle/2))**2  # Passive
        
        # Soil resistance calculations
        if pgd_path == 'Parallel':
            # Longitudinal resistance (axial soil springs)
            # Based on API RP 1111 and similar standards
            friction_coeff = math.tan(math.radians(friction_angle))
            normal_stress = unit_weight * pipe_doc  # Overburden stress
            
            # Adhesion factor for cohesive soils
            if cohesion > 0:
                adhesion_factor = min(1.0, 0.5 * (1 + cohesion/1000))
            else:
                adhesion_factor = 1.0
            
            # Longitudinal force per unit length (lb/ft)
            longitudinal_force = (friction_coeff * normal_stress + adhesion_factor * cohesion) * 3.14159 * pipe_od_ft
            
        else:  # Perpendicular
            # Transverse resistance (lateral soil springs)
            # Based on lateral earth pressure theory
            
            # Net ultimate lateral resistance
            lateral_resistance_cohesion = cohesion * pipe_od_ft * 9.0  # Nc factor ~9 for deep foundations
            lateral_resistance_friction = 0.5 * unit_weight * pipe_doc**2 * kp * pipe_od_ft
            
            longitudinal_force = lateral_resistance_cohesion + lateral_resistance_friction
        
        # Stress calculations
        # Axial stress from longitudinal force
        axial_stress = longitudinal_force * pipe_length / pipe_area_sqin
        
        # Hoop stress from internal pressure
        hoop_stress = internal_pressure * inner_radius / pipe_wt
        
        # Allowable stress (typically 72% of SMYS for combined loading)
        allowable_stress = 0.72 * smys_psi
        
        # Total applied stress (axial + hoop, simplified approach)
        total_applied_stress = axial_stress + 0.5 * hoop_stress  # Conservative combination
        
        # Remaining allowable stress
        remaining_allowable = max(0, allowable_stress - total_applied_stress)
        
        # Allowable pipe length calculation
        if longitudinal_force > 0:
            # Maximum allowable axial stress considering hoop stress
            max_allowable_axial = allowable_stress - 0.5 * hoop_stress
            if max_allowable_axial > 0:
                allowable_length = max_allowable_axial * pipe_area_sqin / longitudinal_force
            else:
                allowable_length = 0
        else:
            allowable_length = 1000  # Large value if no force
        
        # Exceeds allowable check - compare total stress to allowable
        exceeds_allowable = "Exceeds" if total_applied_stress > allowable_stress else "Does Not Exceed"
        
        # Return results in the format matching Excel output headers
        return {
            # Input section (B3:B9 headers, C3:C9 values)
            'Pipe OD (in)': pipe_od,
            'Pipe wt (in)': pipe_wt,
            'Pipe SMYS (psi)': pipe_smys,
            'Pipe DOC (ft)': pipe_doc,
            'Length of Pipe in PGD (ft)': pipe_length,
            'Pipe Coating': pipe_coating,
            'Internal Pressure (psi)': internal_pressure,
            
            # Output section (B13:B17 headers, C13:C17 values)
            'Longitudinal Force (lb/ft)': round(longitudinal_force, 2),
            'Axial Stress (psi)': round(axial_stress, 2),
            'Remaining Allowable Stress (psi)': round(remaining_allowable, 2),
            'Allowable Pipe Length in PGD (ft)': round(allowable_length, 2),
            'Exceeds Allowable': exceeds_allowable,
            
            # Soil section (E3:E6 headers, F3:F6 values)
            'Soil Friction Angle (φ degrees)': friction_angle,
            'Soil Cohesion (c, psf)': cohesion,
            "Soil Effective Unit Weight (γ', psf)": unit_weight,
            'PGD Path (perpendicular/parallel to pipe)': pgd_path
        }
    
    def generate_combinations_with_calculations(self, soil_layer: Dict) -> List[Dict[str, Any]]:
        """Generate parameter combinations with calculated results for a soil layer."""
        logger.info(f"Generating combinations with calculations for soil layer: {soil_layer['name']}")
        
        pipe_ranges = self.generate_pipe_parameter_ranges()
        param_names = list(pipe_ranges.keys())
        param_values = [pipe_ranges[name] for name in param_names]
        
        total_combinations = 1
        for values in param_values:
            total_combinations *= len(values)
        
        logger.info(f"Processing {total_combinations} combinations...")
        
        combinations = []
        combination_count = 0
        
        for combo in itertools.product(*param_values):
            combination_count += 1
            if combination_count % 500 == 0:
                logger.info(f"Processing combination {combination_count}/{total_combinations}")
            
            pipe_params = dict(zip(param_names, combo))
            
            # Calculate results using manual formulas
            calculated_results = self.calculate_soil_springs_results(pipe_params, soil_layer)
            
            # Add soil layer info
            result_row = {
                'Soil Name': soil_layer['name'],
                'Soil Type': soil_layer['type']
            }
            
            # Add all calculated results
            result_row.update(calculated_results)
            
            combinations.append(result_row)
        
        logger.info(f"Generated {len(combinations)} combinations with calculations")
        return combinations
    
    def save_combinations_to_csv(self, combinations: List[Dict[str, Any]], soil_layer_name: str, 
                                output_dir: str = "efficient_static_values_output") -> str:
        """Save parameter combinations with calculated results to CSV file."""
        if not combinations:
            logger.warning(f"No combinations to save for {soil_layer_name}")
            return ""
        
        os.makedirs(output_dir, exist_ok=True)
        
        safe_name = soil_layer_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        csv_filename = os.path.join(output_dir, f"{safe_name}_calculations.csv")
        
        # Define column order
        column_order = [
            'Soil Name', 'Soil Type',
            # Input parameters (matching Excel B3:B9)
            'Pipe OD (in)', 'Pipe wt (in)', 'Pipe SMYS (psi)', 'Pipe DOC (ft)', 
            'Length of Pipe in PGD (ft)', 'Pipe Coating', 'Internal Pressure (psi)',
            # Soil parameters (matching Excel E3:E6)
            'Soil Friction Angle (φ degrees)', 'Soil Cohesion (c, psf)', 
            "Soil Effective Unit Weight (γ', psf)", 'PGD Path (perpendicular/parallel to pipe)',
            # Output results (matching Excel B13:B17)
            'Longitudinal Force (lb/ft)', 'Axial Stress (psi)', 'Remaining Allowable Stress (psi)', 
            'Allowable Pipe Length in PGD (ft)', 'Exceeds Allowable'
        ]
        
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=column_order)
            writer.writeheader()
            writer.writerows(combinations)
        
        logger.info(f"Saved {len(combinations)} combinations to {csv_filename}")
        return csv_filename
    
    def run_complete_analysis(self, output_dir: str = "efficient_static_values_output") -> List[str]:
        """Run the complete analysis with manual calculations."""
        logger.info("Starting efficient static values analysis with manual calculations...")
        
        self.load_pipe_assumptions()
        self.load_soil_assumptions()
        
        generated_files = []
        
        for i, soil_layer in enumerate(self.soil_layers, 1):
            logger.info(f"Processing soil layer {i}/{len(self.soil_layers)}: {soil_layer['name']}")
            
            combinations = self.generate_combinations_with_calculations(soil_layer)
            csv_file = self.save_combinations_to_csv(combinations, soil_layer['name'], output_dir)
            
            if csv_file:
                generated_files.append(csv_file)
        
        logger.info(f"Efficient analysis complete! Generated {len(generated_files)} CSV files")
        return generated_files


def main():
    """Main execution function."""
    calculator = EfficientStaticValuesCalculator()
    
    # Load and display summary
    calculator.load_pipe_assumptions()
    calculator.load_soil_assumptions()
    
    print("\n" + "="*60)
    print("EFFICIENT STATIC VALUES CALCULATOR")
    print("="*60)
    print(f"📁 Static Values File: {calculator.static_values_path}")
    print(f"🏔️ Soil Layers: {len(calculator.soil_layers)}")
    
    pipe_ranges = calculator.generate_pipe_parameter_ranges()
    total_combinations = 1
    for values in pipe_ranges.values():
        total_combinations *= len(values)
    
    print(f"📊 Total combinations per soil layer: {total_combinations:,}")
    print(f"💻 Calculation Method: Manual engineering formulas (fast)")
    
    # Run analysis
    print(f"\n🚀 Starting CSV generation with manual calculations...")
    generated_files = calculator.run_complete_analysis()
    
    print(f"\n✅ EFFICIENT ANALYSIS COMPLETE!")
    print(f"Generated {len(generated_files)} CSV files with calculated results:")
    for file_path in generated_files:
        print(f"  📄 {file_path}")


if __name__ == "__main__":
    main()