#!/usr/bin/env python3
"""
Enhanced Static Values Iterator with Soil Springs Integration
Extracts pipe and soil assumptions from Static Values.xlsx, applies them to
Soil Springs_2024.xlsx for calculations, and generates CSV files with results.
"""

import openpyxl
import pandas as pd
import itertools
import csv
import os
import shutil
from typing import List, Dict, Tuple, Any
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class EnhancedStaticValuesIterator:
    def __init__(self, static_values_path: str = "Static Values.xlsx", 
                 soil_springs_path: str = "Soil Springs_2024.xlsx"):
        """Initialize the iterator with Excel file paths."""
        self.static_values_path = static_values_path
        self.soil_springs_path = soil_springs_path
        self.pipe_assumptions = {}
        self.soil_layers = []
        
        # Define cell mappings for Soil Springs_2024.xlsx
        self.input_cells = {
            'Pipe OD (in)': 'C3',
            'Pipe wt (in)': 'C4',
            'Pipe SMYS (psi)': 'C5',
            'Pipe DOC (ft)': 'C6',
            'Length of Pipe in PGD (ft)': 'C7',
            'Pipe Coating': 'C8',
            'Internal Pressure (psi)': 'C9',
            'Soil Friction Angle (φ degrees)': 'F3',
            'Soil Cohesion (c, psf)': 'F4',
            "Soil Effective Unit Weight (γ', psf)": 'F5',
            'PGD Path (perpendicular/parallel to pipe)': 'F6'
        }
        
        # Define output cell ranges and their headers
        self.output_mappings = {
            # Headers from B3:B9, values from C3:C9
            'input_section': {
                'header_range': [(3, 2), (4, 2), (5, 2), (6, 2), (7, 2), (8, 2), (9, 2)],
                'value_range': [(3, 3), (4, 3), (5, 3), (6, 3), (7, 3), (8, 3), (9, 3)]
            },
            # Headers from B13:B17, values from C13:C17
            'output_section': {
                'header_range': [(13, 2), (14, 2), (15, 2), (16, 2), (17, 2)],
                'value_range': [(13, 3), (14, 3), (15, 3), (16, 3), (17, 3)]
            },
            # Headers from E3:E6, values from F3:F6
            'soil_section': {
                'header_range': [(3, 5), (4, 5), (5, 5), (6, 5)],
                'value_range': [(3, 6), (4, 6), (5, 6), (6, 6)]
            }
        }
    
    def load_pipe_assumptions(self) -> Dict[str, Dict[str, Any]]:
        """Load pipe assumptions from the Static Values Excel file."""
        logger.info("Loading pipe assumptions...")
        
        wb = openpyxl.load_workbook(self.static_values_path, data_only=True)
        ws = wb['Pipe Assumptions']
        
        pipe_assumptions = {}
        
        for row in range(2, ws.max_row + 1):
            param_name = ws.cell(row=row, column=1).value
            min_value = ws.cell(row=row, column=2).value
            max_value = ws.cell(row=row, column=3).value
            
            if param_name and param_name.strip() and min_value is not None:
                param_name = param_name.strip().rstrip(':')
                
                if isinstance(min_value, (int, float)) and isinstance(max_value, (int, float)):
                    pipe_assumptions[param_name] = {
                        'min': min_value,
                        'max': max_value,
                        'type': 'numeric'
                    }
                else:
                    pipe_assumptions[param_name] = {
                        'min': str(min_value),
                        'max': str(max_value),
                        'type': 'categorical'
                    }
        
        wb.close()
        self.pipe_assumptions = pipe_assumptions
        logger.info(f"Loaded {len(pipe_assumptions)} pipe parameters")
        return pipe_assumptions
    
    def load_soil_assumptions(self) -> List[Dict[str, Any]]:
        """Load soil layer assumptions from the Static Values Excel file."""
        logger.info("Loading soil assumptions...")
        
        wb = openpyxl.load_workbook(self.static_values_path, data_only=True)
        ws = wb['Soil Assumptions']
        
        soil_layers = []
        
        for row in range(2, ws.max_row + 1):
            soil_name = ws.cell(row=row, column=1).value
            soil_type = ws.cell(row=row, column=2).value
            unit_weight = ws.cell(row=row, column=3).value
            cohesion = ws.cell(row=row, column=4).value
            friction_angle = ws.cell(row=row, column=5).value
            
            if soil_name and soil_name.strip():
                soil_layer = {
                    'name': soil_name.strip(),
                    'type': soil_type.strip() if soil_type else '',
                    'unit_weight': float(unit_weight) if unit_weight else 120.0,
                    'cohesion': float(cohesion) if cohesion else 0.0,
                    'friction_angle': float(friction_angle) if friction_angle else 30.0
                }
                soil_layers.append(soil_layer)
        
        wb.close()
        self.soil_layers = soil_layers
        logger.info(f"Loaded {len(soil_layers)} soil layers")
        return soil_layers
    
    def generate_pipe_parameter_ranges(self) -> Dict[str, List[Any]]:
        """Generate parameter ranges for pipe assumptions."""
        logger.info("Generating pipe parameter ranges...")
        
        parameter_ranges = {}
        
        for param_name, param_data in self.pipe_assumptions.items():
            if param_data['type'] == 'numeric':
                min_val = param_data['min']
                max_val = param_data['max']
                
                if 'DOC' in param_name or 'Length' in param_name:
                    # Generate 1-foot increments
                    values = list(range(int(min_val), int(max_val) + 1, 1))
                    parameter_ranges[param_name] = values
                else:
                    if min_val == max_val:
                        parameter_ranges[param_name] = [min_val]
                    else:
                        parameter_ranges[param_name] = [min_val, max_val]
            else:
                if param_data['min'] == param_data['max']:
                    parameter_ranges[param_name] = [param_data['min']]
                else:
                    parameter_ranges[param_name] = [param_data['min'], param_data['max']]
        
        return parameter_ranges
    
    def extract_headers_and_values_from_soil_springs(self, temp_file_path: str) -> Tuple[List[str], List[Any]]:
        """Extract headers and calculated values from Soil Springs Excel file."""
        wb = openpyxl.load_workbook(temp_file_path, data_only=True)
        ws = wb['Input&Summary']
        
        headers = []
        values = []
        
        # Extract from each section
        for section_name, section_data in self.output_mappings.items():
            for (header_row, header_col), (value_row, value_col) in zip(
                section_data['header_range'], section_data['value_range']):
                
                header = ws.cell(row=header_row, column=header_col).value
                value = ws.cell(row=value_row, column=value_col).value
                
                if header:
                    header = str(header).strip().rstrip(':')
                    headers.append(header)
                    values.append(value)
        
        wb.close()
        return headers, values
    
    def update_soil_springs_parameters(self, temp_file_path: str, pipe_params: Dict, soil_layer: Dict) -> None:
        """Update parameters in the Soil Springs Excel file."""
        wb = openpyxl.load_workbook(temp_file_path)
        ws = wb['Input&Summary']
        
        # Update pipe parameters
        for param_name, value in pipe_params.items():
            cell = self.input_cells.get(param_name)
            if cell:
                ws[cell] = value
        
        # Update soil parameters
        ws['F3'] = soil_layer['friction_angle']  # Soil Friction Angle
        ws['F4'] = soil_layer['cohesion']        # Soil Cohesion
        ws['F5'] = soil_layer['unit_weight']     # Soil Unit Weight
        
        wb.save(temp_file_path)
        wb.close()
    
    def calculate_with_excel(self, pipe_params: Dict, soil_layer: Dict) -> Tuple[List[str], List[Any]]:
        """
        Calculate results using Excel formulas.
        This method requires Excel to be installed for formula calculation.
        """
        try:
            import xlwings as xw
            
            # Create temporary copy of Soil Springs file
            temp_file = f"temp_soil_springs_{os.getpid()}.xlsx"
            shutil.copy2(self.soil_springs_path, temp_file)
            
            # Update parameters and calculate
            with xw.App(visible=False, add_book=False) as app:
                wb = app.books.open(os.path.abspath(temp_file))
                ws = wb.sheets['Input&Summary']
                
                # Update pipe parameters
                for param_name, value in pipe_params.items():
                    cell = self.input_cells.get(param_name)
                    if cell:
                        ws.range(cell).value = value
                
                # Update soil parameters
                ws.range('F3').value = soil_layer['friction_angle']
                ws.range('F4').value = soil_layer['cohesion']
                ws.range('F5').value = soil_layer['unit_weight']
                
                # Force calculation
                app.calculate()
                
                # Extract results
                headers, values = self.extract_headers_and_values_from_soil_springs(temp_file)
                
                wb.close()
            
            # Clean up
            if os.path.exists(temp_file):
                os.remove(temp_file)
            
            return headers, values
            
        except ImportError:
            logger.warning("xlwings not available, using openpyxl fallback")
            return self.calculate_with_openpyxl(pipe_params, soil_layer)
        except Exception as e:
            logger.error(f"Excel calculation failed: {e}")
            return self.calculate_with_openpyxl(pipe_params, soil_layer)
    
    def calculate_with_openpyxl(self, pipe_params: Dict, soil_layer: Dict) -> Tuple[List[str], List[Any]]:
        """
        Fallback calculation using openpyxl (formulas won't be calculated).
        This method updates the parameters but can't recalculate Excel formulas.
        """
        # Create temporary copy
        temp_file = f"temp_soil_springs_{os.getpid()}.xlsx"
        shutil.copy2(self.soil_springs_path, temp_file)
        
        # Update parameters
        self.update_soil_springs_parameters(temp_file, pipe_params, soil_layer)
        
        # Extract headers and current values (won't be recalculated)
        headers, values = self.extract_headers_and_values_from_soil_springs(temp_file)
        
        # Clean up
        if os.path.exists(temp_file):
            os.remove(temp_file)
        
        logger.warning("Using static values - formulas not recalculated (Excel/xlwings required for calculation)")
        return headers, values
    
    def generate_combinations_with_calculations(self, soil_layer: Dict) -> List[Dict[str, Any]]:
        """Generate parameter combinations with calculated results for a soil layer."""
        logger.info(f"Generating combinations with calculations for soil layer: {soil_layer['name']}")
        
        pipe_ranges = self.generate_pipe_parameter_ranges()
        param_names = list(pipe_ranges.keys())
        param_values = [pipe_ranges[name] for name in param_names]
        
        combinations = []
        total_combinations = 1
        for values in param_values:
            total_combinations *= len(values)
        
        logger.info(f"Processing {total_combinations} combinations...")
        
        # Get headers from first calculation
        first_combo = dict(zip(param_names, [values[0] for values in param_values]))
        headers, _ = self.calculate_with_excel(first_combo, soil_layer)
        
        combination_count = 0
        for combo in itertools.product(*param_values):
            combination_count += 1
            if combination_count % 100 == 0:
                logger.info(f"Processing combination {combination_count}/{total_combinations}")
            
            pipe_params = dict(zip(param_names, combo))
            
            # Calculate results
            _, calculated_values = self.calculate_with_excel(pipe_params, soil_layer)
            
            # Combine parameters and results
            result_row = {}
            
            # Add soil layer info first
            result_row.update({
                'Soil Name': soil_layer['name'],
                'Soil Type': soil_layer['type']
            })
            
            # Add calculated results with headers
            for header, value in zip(headers, calculated_values):
                result_row[header] = value
            
            combinations.append(result_row)
        
        logger.info(f"Generated {len(combinations)} combinations with calculations")
        return combinations
    
    def save_combinations_to_csv(self, combinations: List[Dict[str, Any]], soil_layer_name: str, 
                                output_dir: str = "enhanced_static_values_output") -> str:
        """Save parameter combinations with calculated results to CSV file."""
        if not combinations:
            logger.warning(f"No combinations to save for {soil_layer_name}")
            return ""
        
        os.makedirs(output_dir, exist_ok=True)
        
        safe_name = soil_layer_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        csv_filename = os.path.join(output_dir, f"{safe_name}_calculations.csv")
        
        # Get all column names
        all_columns = list(combinations[0].keys())
        
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=all_columns)
            writer.writeheader()
            writer.writerows(combinations)
        
        logger.info(f"Saved {len(combinations)} combinations to {csv_filename}")
        return csv_filename
    
    def run_complete_analysis(self, output_dir: str = "enhanced_static_values_output") -> List[str]:
        """Run the complete analysis with Excel calculations."""
        logger.info("Starting enhanced static values analysis with Excel calculations...")
        
        self.load_pipe_assumptions()
        self.load_soil_assumptions()
        
        generated_files = []
        
        for i, soil_layer in enumerate(self.soil_layers, 1):
            logger.info(f"Processing soil layer {i}/{len(self.soil_layers)}: {soil_layer['name']}")
            
            combinations = self.generate_combinations_with_calculations(soil_layer)
            csv_file = self.save_combinations_to_csv(combinations, soil_layer['name'], output_dir)
            
            if csv_file:
                generated_files.append(csv_file)
        
        logger.info(f"Enhanced analysis complete! Generated {len(generated_files)} CSV files")
        return generated_files


def main():
    """Main execution function."""
    iterator = EnhancedStaticValuesIterator()
    
    # Load and display summary
    iterator.load_pipe_assumptions()
    iterator.load_soil_assumptions()
    
    print("\n" + "="*60)
    print("ENHANCED STATIC VALUES ANALYSIS")
    print("="*60)
    print(f"📁 Static Values File: {iterator.static_values_path}")
    print(f"🧮 Soil Springs File: {iterator.soil_springs_path}")
    print(f"🏔️ Soil Layers: {len(iterator.soil_layers)}")
    
    pipe_ranges = iterator.generate_pipe_parameter_ranges()
    total_combinations = 1
    for values in pipe_ranges.values():
        total_combinations *= len(values)
    
    print(f"📊 Total combinations per soil layer: {total_combinations:,}")
    print(f"💻 Calculation Method: Excel formulas (xlwings) with openpyxl fallback")
    
    # Run analysis
    print(f"\n🚀 Starting CSV generation with calculations...")
    generated_files = iterator.run_complete_analysis()
    
    print(f"\n✅ ENHANCED ANALYSIS COMPLETE!")
    print(f"Generated {len(generated_files)} CSV files with calculated results:")
    for file_path in generated_files:
        print(f"  📄 {file_path}")


if __name__ == "__main__":
    main()